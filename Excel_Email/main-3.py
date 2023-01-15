import requests # 사이트 접속을 위한 모듈ㄴ
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url = "https://www.hani.co.kr/arti/society/society_general/1075274.html";

headers = {
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'text/html; charset=utf-8'
} # 헤더정보를 입력하지 않을시 사이트에서 로봇이 접속한 것으로 판단하여 응답하지 않습니다.

response = requests.get(url, headers = headers);

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text);

results = list(set(results));

print(results);

try:
    wb = load_workbook('Email.xlsx', data_only=True); # Email.xlsx 파일이 있다면
    sheet = wb.active;                                # 파일을 읽습니다.
except:
    wb = Workbook(); # Email.xlsx 파일이 없다면 새로운 엑셀을 생성합니다.
    sheet = wb.active;

for result in results:
    sheet.append([result]); # 이메일을 수집한 수만큼 엑셀에 기록합니다.

wb.save('Email.xlsx');
